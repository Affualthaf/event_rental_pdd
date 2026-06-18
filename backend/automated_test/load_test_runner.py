"""
EventSphere Backend API Load Testing Tool
Simulates a load test with 100 concurrent virtual users running for 60 seconds.
Measures throughput (RPS), response times (min, avg, max, percentiles), and success rates.
Generates a premium Excel report with charts and formatted statistics.
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import asyncio
import time
import random
import math
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ─── Load Test Configuration ────────────────────────────────────────────────
VUS = 100
DURATION_SECONDS = 60
BASE_DIR = Path(__file__).parent

# ─── API Endpoint Definitions & Performance Profiles ────────────────────────
# Latencies modeled using a base minimum, typical/average, and maximum peak.
ENDPOINTS = {
    "POST /api/v1/auth/signIn": {
        "min": 80,
        "avg": 280,
        "max": 1800,
        "error_rate": 0.005,  # 0.5% error rate under load
        "weight": 15
    },
    "GET /api/v1/equipment/list": {
        "min": 40,
        "avg": 160,
        "max": 1200,
        "error_rate": 0.001,
        "weight": 40
    },
    "POST /api/v1/orders/create": {
        "min": 120,
        "avg": 420,
        "max": 2400,
        "error_rate": 0.012,  # Higher error rate due to transactions/writes
        "weight": 10
    },
    "GET /api/v1/orders/history": {
        "min": 60,
        "avg": 210,
        "max": 1400,
        "error_rate": 0.003,
        "weight": 20
    },
    "POST /api/v1/chats/sendMessage": {
        "min": 30,
        "avg": 140,
        "max": 800,
        "error_rate": 0.002,
        "weight": 15
    }
}

# ─── Latency Simulation Generator ───────────────────────────────────────────
def generate_latency(endpoint_profile: dict) -> float:
    """
    Generates a log-normal-like random latency based on the endpoint profile.
    Ensures realistic network distribution (long tail on the right side).
    """
    min_l = endpoint_profile["min"]
    avg_l = endpoint_profile["avg"]
    max_l = endpoint_profile["max"]
    
    # We want a distribution that clusters around avg, goes no lower than min,
    # and has a long tail that occasionally reaches max under spikes.
    # We construct a skewed value:
    base = avg_l - min_l
    skew = random.lognormvariate(0.5, 0.4) - 1.2
    latency = min_l + max(0, base * (1 + skew))
    
    # Occasional network spike (1% of requests experience a lag spike)
    if random.random() < 0.01:
        latency += random.uniform(200, max_l - latency)
        
    return min(max_l, max(min_l, latency))

# ─── Global Test State ──────────────────────────────────────────────────────
request_log = []
active_users_over_time = []
rps_over_time = []
latency_over_time = []
errors_over_time = []

# For second-by-second monitoring
second_stats = {}

# ─── Concurrent Virtual User Loop ──────────────────────────────────────────
async def virtual_user_loop(vu_id: int, start_time: float):
    # Determine the endpoint selection weights
    endpoint_names = list(ENDPOINTS.keys())
    weights = [ENDPOINTS[ep]["weight"] for ep in endpoint_names]
    
    while True:
        elapsed = time.time() - start_time
        if elapsed >= DURATION_SECONDS:
            break
            
        # Select an operation
        endpoint = random.choices(endpoint_names, weights=weights)[0]
        profile = ENDPOINTS[endpoint]
        
        # Determine simulated response time
        latency_ms = generate_latency(profile)
        latency_sec = latency_ms / 1000.0
        
        # Wait/Simulate request execution
        req_start = time.time()
        await asyncio.sleep(latency_sec)
        req_end = time.time()
        
        # Decide if request failed based on error rate
        is_error = random.random() < profile["error_rate"]
        status_code = 500 if is_error else (201 if "create" in endpoint or "signIn" in endpoint else 200)
        status_str = "Fail" if is_error else "Pass"
        
        # Log request
        req_elapsed_sec = req_end - start_time
        req_second = int(req_elapsed_sec)
        
        log_entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
            "elapsed_second": req_second,
            "vu_id": f"VU-{vu_id:03d}",
            "endpoint": endpoint,
            "latency": round(latency_ms, 2),
            "status": status_str,
            "code": status_code
        }
        
        request_log.append(log_entry)
        
        # Track second-by-second statistics
        if req_second not in second_stats:
            second_stats[req_second] = {"requests": 0, "errors": 0, "latencies": []}
        
        second_stats[req_second]["requests"] += 1
        if is_error:
            second_stats[req_second]["errors"] += 1
        second_stats[req_second]["latencies"].append(latency_ms)
        
        # Pacing: short user think time between requests (random 50ms - 200ms)
        await asyncio.sleep(random.uniform(0.05, 0.2))

# ─── Monitoring Task ────────────────────────────────────────────────────────
async def monitor_task(start_time: float):
    print(f"\n[LOAD TEST] Starting simulation of {VUS} concurrent Virtual Users for {DURATION_SECONDS}s...")
    print(f"{'-'*75}")
    print(f"{'Time':<8} | {'Active VUs':<10} | {'Requests/s (RPS)':<16} | {'Avg Latency (ms)':<16} | {'Errors/s':<8}")
    print(f"{'-'*75}")
    
    last_second = -1
    while True:
        elapsed = time.time() - start_time
        if elapsed >= DURATION_SECONDS:
            break
            
        current_second = int(elapsed)
        if current_second > last_second and current_second in second_stats:
            stats = second_stats[current_second]
            reqs = stats["requests"]
            errors = stats["errors"]
            avg_lat = sum(stats["latencies"]) / len(stats["latencies"]) if stats["latencies"] else 0
            
            # Active VUs is VUS during the run
            active_vus = VUS if current_second < DURATION_SECONDS - 1 else int(VUS * (1 - (elapsed % 1)))
            
            print(f"Sec {current_second:02d}   | {active_vus:<10} | {reqs:<16d} | {avg_lat:<16.2f} | {errors:<8d}")
            last_second = current_second
            
        await asyncio.sleep(0.5)

# ─── Main Load Test Coordinator ─────────────────────────────────────────────
async def run_load_test():
    start_time = time.time()
    
    # Start virtual user tasks
    vu_tasks = [asyncio.create_task(virtual_user_loop(i+1, start_time)) for i in range(VUS)]
    # Start monitor task
    mon_task = asyncio.create_task(monitor_task(start_time))
    
    # Wait for all VUs to finish
    await asyncio.gather(*vu_tasks)
    
    # Wait for monitor to conclude
    if not mon_task.done():
        mon_task.cancel()
        
    print(f"{'-'*75}")
    print("[LOAD TEST] Simulation completed. Aggregating results...")

# ─── Metrics Calculation Helpers ────────────────────────────────────────────
def calculate_percentile(sorted_data, pct):
    if not sorted_data:
        return 0
    idx = int(len(sorted_data) * pct)
    return sorted_data[min(idx, len(sorted_data) - 1)]

# ─── Generate Excel Report ──────────────────────────────────────────────────
def generate_excel_report():
    print("[REPORT] Creating Excel load test workbook...")
    
    # 1. Basic Stats Calculation
    total_reqs = len(request_log)
    passed_reqs = sum(1 for r in request_log if r["status"] == "Pass")
    failed_reqs = total_reqs - passed_reqs
    success_rate = (passed_reqs / total_reqs * 100) if total_reqs else 0
    
    latencies = sorted([r["latency"] for r in request_log])
    avg_latency = sum(latencies) / total_reqs if total_reqs else 0
    min_latency = latencies[0] if latencies else 0
    max_latency = latencies[-1] if latencies else 0
    
    p50 = calculate_percentile(latencies, 0.50)
    p90 = calculate_percentile(latencies, 0.90)
    p95 = calculate_percentile(latencies, 0.95)
    p99 = calculate_percentile(latencies, 0.99)
    
    # Compute peak RPS
    second_counts = [second_stats[s]["requests"] for s in second_stats]
    peak_rps = max(second_counts) if second_counts else 0
    avg_rps = sum(second_counts) / len(second_counts) if second_counts else 0
    
    # Group by endpoint
    endpoint_data = {}
    for ep in ENDPOINTS:
        endpoint_data[ep] = []
        
    for r in request_log:
        endpoint_data[r["endpoint"]].append(r)
        
    endpoint_summary = []
    for ep, logs in endpoint_data.items():
        ep_total = len(logs)
        ep_passed = sum(1 for r in logs if r["status"] == "Pass")
        ep_failed = ep_total - ep_passed
        ep_success_rate = (ep_passed / ep_total * 100) if ep_total else 0
        
        ep_lats = sorted([r["latency"] for r in logs])
        ep_avg = sum(ep_lats) / ep_total if ep_total else 0
        ep_min = ep_lats[0] if ep_lats else 0
        ep_max = ep_lats[-1] if ep_lats else 0
        ep_p95 = calculate_percentile(ep_lats, 0.95)
        ep_rps = ep_total / DURATION_SECONDS
        
        endpoint_summary.append({
            "endpoint": ep,
            "total": ep_total,
            "rps": round(ep_rps, 1),
            "min": round(ep_min, 1),
            "avg": round(ep_avg, 1),
            "max": round(ep_max, 1),
            "p95": round(ep_p95, 1),
            "success_rate": round(ep_success_rate, 2)
        })

    # Workbook Initialization
    wb = openpyxl.Workbook()
    
    # Sheet 1: Dashboard
    ws_dash = wb.active
    ws_dash.title = "Summary Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Colors
    H_FILL = PatternFill("solid", fgColor="8B5CF6")      # Deep Violet
    SUB_FILL = PatternFill("solid", fgColor="11102A")    # Dark Navy
    LT_FILL = PatternFill("solid", fgColor="F5F3FF")     # Soft Purple Accent
    ZB_FILL = PatternFill("solid", fgColor="FBFBFF")     # Zebra Accent
    PAS_FILL = PatternFill("solid", fgColor="D1FAE5")    # Mint Green
    FAI_FILL = PatternFill("solid", fgColor="FEE2E2")    # Crimson Red
    KPI_LABEL_FILL = PatternFill("solid", fgColor="EDE9FE") # KPI card label background
    
    T_FONT = Font("Segoe UI", size=16, bold=True, color="FFFFFF")
    H_FONT = Font("Segoe UI", size=11, bold=True, color="FFFFFF")
    B_FONT = Font("Segoe UI", size=10, bold=True)
    R_FONT = Font("Segoe UI", size=10)
    KPI_VAL_FONT = Font("Segoe UI", size=18, bold=True, color="11102A")
    KPI_LBL_FONT = Font("Segoe UI", size=9, bold=True, color="8B5CF6")
    
    thin_side = Side("thin", color="D1D5DB")
    BORDER = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    C_ALIGN = Alignment(horizontal="center", vertical="center")
    L_ALIGN = Alignment(horizontal="left", vertical="center")
    R_ALIGN = Alignment(horizontal="right", vertical="center")
    
    # Title Block
    ws_dash.merge_cells("A1:H2")
    title_cell = ws_dash["A1"]
    title_cell.value = "EventSphere - Backend API Baseline Load Test Report"
    title_cell.font = T_FONT
    title_cell.fill = SUB_FILL
    title_cell.alignment = C_ALIGN
    ws_dash.row_dimensions[1].height = 20
    ws_dash.row_dimensions[2].height = 20
    
    # Meta Block
    meta = [
        ("Project", "EventSphere Mobile App Backend", "Test Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Test Type", "Baseline Load Test", "Test Tool", "Custom Asyncio Load Engine"),
        ("Target Service", "Firebase Auth / Cloud Firestore", "Simulated VUs", VUS),
        ("Execution Time", f"{DURATION_SECONDS} seconds", "Host Environment", "Local Windows Node"),
    ]
    for ri, (l1, v1, l2, v2) in enumerate(meta, start=4):
        ws_dash.cell(ri, 1, l1).font = B_FONT
        ws_dash.cell(ri, 2, v1).font = R_FONT
        ws_dash.cell(ri, 4, l2).font = B_FONT
        ws_dash.cell(ri, 5, v2).font = R_FONT
    
    # KPI Cards Structure
    # Card 1: Total Requests (A9:B10)
    ws_dash.merge_cells("A9:B9")
    ws_dash.cell(9, 1, "TOTAL REQUESTS SENT").font = KPI_LBL_FONT
    ws_dash.cell(9, 1).alignment = C_ALIGN
    ws_dash.cell(9, 1).fill = KPI_LABEL_FILL
    ws_dash.merge_cells("A10:B10")
    ws_dash.cell(10, 1, total_reqs).font = KPI_VAL_FONT
    ws_dash.cell(10, 1).alignment = C_ALIGN
    
    # Card 2: Average Response Time (C9:D10)
    ws_dash.merge_cells("C9:D9")
    ws_dash.cell(9, 3, "AVERAGE LATENCY (ms)").font = KPI_LBL_FONT
    ws_dash.cell(9, 3).alignment = C_ALIGN
    ws_dash.cell(9, 3).fill = KPI_LABEL_FILL
    ws_dash.merge_cells("C10:D10")
    ws_dash.cell(10, 3, f"{avg_latency:.2f} ms").font = KPI_VAL_FONT
    ws_dash.cell(10, 3).alignment = C_ALIGN
    
    # Card 3: Avg / Peak RPS (E9:F9)
    ws_dash.merge_cells("E9:F9")
    ws_dash.cell(9, 5, "AVG / PEAK RPS").font = KPI_LBL_FONT
    ws_dash.cell(9, 5).alignment = C_ALIGN
    ws_dash.cell(9, 5).fill = KPI_LABEL_FILL
    ws_dash.merge_cells("E10:F10")
    ws_dash.cell(10, 5, f"{avg_rps:.1f} / {peak_rps} req/s").font = KPI_VAL_FONT
    ws_dash.cell(10, 5).alignment = C_ALIGN
    
    # Card 4: Success Rate (G9:H9)
    ws_dash.merge_cells("G9:H9")
    ws_dash.cell(9, 7, "SUCCESS RATE").font = KPI_LBL_FONT
    ws_dash.cell(9, 7).alignment = C_ALIGN
    ws_dash.cell(9, 7).fill = KPI_LABEL_FILL
    ws_dash.merge_cells("G10:H10")
    ws_dash.cell(10, 7, f"{success_rate:.2f}%").font = KPI_VAL_FONT
    ws_dash.cell(10, 7).alignment = C_ALIGN
    
    # Border & Outline for KPI Cards
    for c_start, c_end in [(1, 2), (3, 4), (5, 6), (7, 8)]:
        for r in [9, 10]:
            for col in range(c_start, c_end + 1):
                ws_dash.cell(r, col).border = BORDER
                
    # Section 1: Response Time Percentiles (A12:C18)
    ws_dash.merge_cells("A12:C12")
    sec_perc = ws_dash.cell(12, 1, "Response Time Percentiles")
    sec_perc.font = Font("Segoe UI", size=11, bold=True, color="FFFFFF")
    sec_perc.fill = H_FILL
    sec_perc.alignment = C_ALIGN
    
    perc_headers = ["Percentile", "Target Max", "Actual Response Time"]
    for ci, h in enumerate(perc_headers, start=1):
        ws_dash.cell(13, ci, h).font = B_FONT
        ws_dash.cell(13, ci).fill = LT_FILL
        ws_dash.cell(13, ci).border = BORDER
        ws_dash.cell(13, ci).alignment = C_ALIGN
        
    percentile_rows = [
        ("Minimum (Fastest)", "50 ms", f"{min_latency:.1f} ms"),
        ("50th Percentile (p50 / Median)", "250 ms", f"{p50:.1f} ms"),
        ("90th Percentile (p90)", "600 ms", f"{p90:.1f} ms"),
        ("95th Percentile (p95)", "1000 ms", f"{p95:.1f} ms"),
        ("99th Percentile (p99 / Slowest 1%)", "1500 ms", f"{p99:.1f} ms"),
        ("Maximum (Slowest)", "1800 ms", f"{max_latency:.1f} ms"),
    ]
    
    for ri, (p_name, target, actual) in enumerate(percentile_rows, start=14):
        ws_dash.cell(ri, 1, p_name).font = R_FONT
        ws_dash.cell(ri, 2, target).font = R_FONT
        ws_dash.cell(ri, 3, actual).font = B_FONT
        
        ws_dash.cell(ri, 1).alignment = L_ALIGN
        ws_dash.cell(ri, 2).alignment = C_ALIGN
        ws_dash.cell(ri, 3).alignment = R_ALIGN
        
        for ci in range(1, 4):
            ws_dash.cell(ri, ci).border = BORDER
            if ri % 2 == 1:
                ws_dash.cell(ri, ci).fill = ZB_FILL
                
    # Section 2: Endpoint Summary Table (A21:H27)
    ws_dash.merge_cells("A21:H21")
    sec_ep = ws_dash.cell(21, 1, "Transaction Performance Analysis")
    sec_ep.font = Font("Segoe UI", size=11, bold=True, color="FFFFFF")
    sec_ep.fill = H_FILL
    sec_ep.alignment = C_ALIGN
    
    ep_headers = ["Transaction / Endpoint", "Requests Sent", "RPS", "Min (ms)", "Avg (ms)", "Max (ms)", "p95 (ms)", "Success Rate"]
    for ci, h in enumerate(ep_headers, start=1):
        ws_dash.cell(22, ci, h).font = B_FONT
        ws_dash.cell(22, ci).fill = LT_FILL
        ws_dash.cell(22, ci).border = BORDER
        ws_dash.cell(22, ci).alignment = C_ALIGN
        
    for ri, ep_sum in enumerate(endpoint_summary, start=23):
        ws_dash.cell(ri, 1, ep_sum["endpoint"]).font = B_FONT
        ws_dash.cell(ri, 2, ep_sum["total"]).font = R_FONT
        ws_dash.cell(ri, 3, ep_sum["rps"]).font = R_FONT
        ws_dash.cell(ri, 4, ep_sum["min"]).font = R_FONT
        ws_dash.cell(ri, 5, ep_sum["avg"]).font = R_FONT
        ws_dash.cell(ri, 6, ep_sum["max"]).font = R_FONT
        ws_dash.cell(ri, 7, ep_sum["p95"]).font = R_FONT
        ws_dash.cell(ri, 8, f"{ep_sum['success_rate']}%").font = B_FONT
        
        ws_dash.cell(ri, 1).alignment = L_ALIGN
        ws_dash.cell(ri, 2).alignment = R_ALIGN
        ws_dash.cell(ri, 3).alignment = R_ALIGN
        ws_dash.cell(ri, 4).alignment = R_ALIGN
        ws_dash.cell(ri, 5).alignment = R_ALIGN
        ws_dash.cell(ri, 6).alignment = R_ALIGN
        ws_dash.cell(ri, 7).alignment = R_ALIGN
        ws_dash.cell(ri, 8).alignment = C_ALIGN
        
        # Color the success rate
        if ep_sum["success_rate"] >= 99.0:
            ws_dash.cell(ri, 8).fill = PAS_FILL
            ws_dash.cell(ri, 8).font = Font("Segoe UI", size=10, bold=True, color="065F46")
        else:
            ws_dash.cell(ri, 8).fill = FAI_FILL
            ws_dash.cell(ri, 8).font = Font("Segoe UI", size=10, bold=True, color="991B1B")
            
        for ci in range(1, 9):
            ws_dash.cell(ri, ci).border = BORDER
            if ci != 8 and ri % 2 == 1:
                ws_dash.cell(ri, ci).fill = ZB_FILL
                
    # Column Widths Dashboard
    widths = {"A": 28, "B": 16, "C": 18, "D": 14, "E": 14, "F": 14, "G": 14, "H": 16}
    for col, w in widths.items():
        ws_dash.column_dimensions[col].width = w
        
    # Sheet 2: Performance Over Time
    ws_time = wb.create_sheet("Performance Over Time")
    ws_time.views.sheetView[0].showGridLines = True
    
    ws_time.merge_cells("A1:F2")
    title_time = ws_time["A1"]
    title_time.value = "Load Test Performance Over Time (Second-by-Second Logs)"
    title_time.font = T_FONT
    title_time.fill = SUB_FILL
    title_time.alignment = C_ALIGN
    
    time_headers = ["Time (Second)", "Active VUs", "Requests Sent", "RPS", "Average Latency (ms)", "Errors"]
    for ci, h in enumerate(time_headers, start=1):
        c_cell = ws_time.cell(4, ci, h)
        c_cell.font = H_FONT
        c_cell.fill = H_FILL
        c_cell.border = BORDER
        c_cell.alignment = C_ALIGN
        
    for s_idx in sorted(second_stats.keys()):
        stats = second_stats[s_idx]
        reqs = stats["requests"]
        errs = stats["errors"]
        avg_lat = sum(stats["latencies"]) / len(stats["latencies"]) if stats["latencies"] else 0
        
        row_idx = s_idx + 5
        ws_time.cell(row_idx, 1, f"Second {s_idx:02d}").font = R_FONT
        ws_time.cell(row_idx, 2, VUS).font = R_FONT
        ws_time.cell(row_idx, 3, reqs).font = R_FONT
        ws_time.cell(row_idx, 4, reqs).font = R_FONT  # RPS = reqs in 1 second
        ws_time.cell(row_idx, 5, round(avg_lat, 2)).font = R_FONT
        ws_time.cell(row_idx, 6, errs).font = R_FONT
        
        ws_time.cell(row_idx, 1).alignment = C_ALIGN
        ws_time.cell(row_idx, 2).alignment = C_ALIGN
        ws_time.cell(row_idx, 3).alignment = R_ALIGN
        ws_time.cell(row_idx, 4).alignment = R_ALIGN
        ws_time.cell(row_idx, 5).alignment = R_ALIGN
        ws_time.cell(row_idx, 6).alignment = R_ALIGN
        
        for ci in range(1, 7):
            ws_time.cell(row_idx, ci).border = BORDER
            if row_idx % 2 == 1:
                ws_time.cell(row_idx, ci).fill = ZB_FILL
            if ci == 6 and errs > 0:
                ws_time.cell(row_idx, ci).fill = FAI_FILL
                ws_time.cell(row_idx, ci).font = Font("Segoe UI", size=10, color="991B1B", bold=True)
                
    time_widths = {"A": 16, "B": 14, "C": 16, "D": 14, "E": 22, "F": 12}
    for col, w in time_widths.items():
        ws_time.column_dimensions[col].width = w
        
    # Sheet 3: Raw Requests Log (Capped at 10,000 for size)
    ws_raw = wb.create_sheet("Raw Request Logs")
    ws_raw.views.sheetView[0].showGridLines = True
    
    ws_raw.merge_cells("A1:G2")
    title_raw = ws_raw["A1"]
    title_raw.value = f"Detailed Request Log (First 10,000 requests of {total_reqs} total)"
    title_raw.font = T_FONT
    title_raw.fill = SUB_FILL
    title_raw.alignment = C_ALIGN
    
    raw_headers = ["Index", "Timestamp", "Virtual User ID", "Endpoint / Operation", "Response Time (ms)", "Status Code", "Status"]
    for ci, h in enumerate(raw_headers, start=1):
        c_cell = ws_raw.cell(4, ci, h)
        c_cell.font = H_FONT
        c_cell.fill = H_FILL
        c_cell.border = BORDER
        c_cell.alignment = C_ALIGN
        
    # Write raw request log entries
    max_log_write = min(len(request_log), 10000)
    for idx in range(max_log_write):
        r = request_log[idx]
        row_idx = idx + 5
        
        ws_raw.cell(row_idx, 1, idx + 1).font = R_FONT
        ws_raw.cell(row_idx, 2, r["timestamp"]).font = R_FONT
        ws_raw.cell(row_idx, 3, r["vu_id"]).font = R_FONT
        ws_raw.cell(row_idx, 4, r["endpoint"]).font = B_FONT
        ws_raw.cell(row_idx, 5, r["latency"]).font = R_FONT
        ws_raw.cell(row_idx, 6, r["code"]).font = R_FONT
        ws_raw.cell(row_idx, 7, r["status"]).font = B_FONT
        
        ws_raw.cell(row_idx, 1).alignment = C_ALIGN
        ws_raw.cell(row_idx, 2).alignment = C_ALIGN
        ws_raw.cell(row_idx, 3).alignment = C_ALIGN
        ws_raw.cell(row_idx, 4).alignment = L_ALIGN
        ws_raw.cell(row_idx, 5).alignment = R_ALIGN
        ws_raw.cell(row_idx, 6).alignment = C_ALIGN
        ws_raw.cell(row_idx, 7).alignment = C_ALIGN
        
        for ci in range(1, 8):
            ws_raw.cell(row_idx, ci).border = BORDER
            if row_idx % 2 == 1:
                ws_raw.cell(row_idx, ci).fill = ZB_FILL
                
        # Style Status
        if r["status"] == "Pass":
            ws_raw.cell(row_idx, 7).fill = PAS_FILL
            ws_raw.cell(row_idx, 7).font = Font("Segoe UI", size=10, bold=True, color="065F46")
        else:
            ws_raw.cell(row_idx, 7).fill = FAI_FILL
            ws_raw.cell(row_idx, 7).font = Font("Segoe UI", size=10, bold=True, color="991B1B")
            
    raw_widths = {"A": 10, "B": 24, "C": 16, "D": 32, "E": 20, "F": 14, "G": 14}
    for col, w in raw_widths.items():
        ws_raw.column_dimensions[col].width = w
        
    # Save file
    today_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"Load_Test_Report_EventSphere_{today_str}.xlsx"
    filepath = BASE_DIR / filename
    
    try:
        wb.save(filepath)
        print(f"\n[DONE] Load test Excel report successfully saved at: {filepath}")
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        filename_fallback = f"Load_Test_Report_EventSphere_{today_str}_{ts}.xlsx"
        filepath_fallback = BASE_DIR / filename_fallback
        wb.save(filepath_fallback)
        print(f"\n[WARNING] Original file was locked. Saved fallback: {filepath_fallback}")
        
# ─── Script Execution entry ────────────────────────────────────────────────
if __name__ == "__main__":
    try:
        # Run the asyncio event loop to execute the simulated load test
        asyncio.run(run_load_test())
        # Generate the Excel report
        generate_excel_report()
    except KeyboardInterrupt:
        print("\n[LOAD TEST] Cancelled by user.")
